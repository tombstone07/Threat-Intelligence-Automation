import os
import time
import requests
import openpyxl
from openpyxl.styles import PatternFill
from tqdm import tqdm
import win32com.client as win32

API_KEY = "1cbad388ce42d2edb9a99d63bb7e9b68f2d8f02239f140a619c54d1131e970e1"
VT_URL = "https://www.virustotal.com/api/v3/ip_addresses/"
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def is_valid_ip(ip):
    parts = ip.strip().split('.')
    if len(parts) != 4:
        return False
    try:
        return all(0 <= int(part) <= 255 for part in parts)
    except ValueError:
        return False

def check_ip_virustotal(ip):
    headers = {"x-apikey": API_KEY}
    response = requests.get(VT_URL + ip, headers=headers)
    if response.status_code == 200:
        data = response.json()
        stats = data['data']['attributes']['last_analysis_stats']
        total_reports = sum(stats.values())
        malicious = stats.get('malicious', 0)
        if total_reports == 0:
            return False
        malicious_ratio = (malicious / total_reports) * 100
        return malicious > 30 and malicious_ratio > 50
    else:
        print(f"Failed to fetch {ip}: Status {response.status_code}")
        return False

def main():
    files = [f for f in os.listdir() if f.endswith(".xlsx")]
    if not files:
        print("No Excel files found!")
        return
    for idx, file in enumerate(files, 1):
        print(f"{idx}. {file}")
    choice = int(input("\nEnter the number of the workbook you want to scan: "))
    selected_file = files[choice - 1]
    password = input("Please enter the password to open the workbook: ")

    print(f"\n🔓 Opening workbook: {selected_file}...")
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    wb = excel.Workbooks.Open(os.path.abspath(selected_file), Password=password)
    tmp_file = "temp_unprotected.xlsx"
    wb.SaveAs(os.path.abspath(tmp_file), Password="")  # Save unprotected
    wb.Close(SaveChanges=False)

    print("🔎 Scanning for IPs...")
    wb_py = openpyxl.load_workbook(tmp_file)
    all_ips = []

    for sheetname in wb_py.sheetnames:
        ws = wb_py[sheetname]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value = cell.value.strip()
                    if is_valid_ip(value):
                        all_ips.append((ws, cell))

    print(f"✅ Found {len(all_ips)} IPs to check.\n")

    for ws, cell in tqdm(all_ips, desc="Checking IPs"):
        ip = cell.value.strip()
        if check_ip_virustotal(ip):
            cell.fill = highlight_fill
        time.sleep(15)

    # Save the updated workbook
    new_file = f"{os.path.splitext(selected_file)[0]}_highlighted.xlsx"
    wb_py.save(new_file)

    print(f"\n🔒 Re-applying password protection to: {new_file}")
    wb2 = excel.Workbooks.Open(os.path.abspath(new_file))
    wb2.SaveAs(os.path.abspath(new_file), Password=password)
    wb2.Close()
    excel.Quit()

    os.remove(tmp_file)
    print(f"\n✅ Finished! Saved and re-encrypted workbook as: {new_file}")

if __name__ == "__main__":
    main()
